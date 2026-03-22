#!/usr/bin/env python3
"""
Parse 'Crime Breakdown' sheet from JC IMPACT weekly Excel workbook and emit ytd-breakdown.json
for the dashboard right-hand YTD panel.

Usage:
  python3 scripts/export_ytd_breakdown.py [path/to/workbook.xlsx]

Default (if no arg): Desktop path used in development — override for CI.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit("Install openpyxl: pip install openpyxl") from e

DEFAULT_XLSX = Path.home() / "Desktop" / "JCIMPACT Weekly Jan Feb 2026.xlsx"
OUT_PATH = Path(__file__).resolve().parent.parent / "ytd-breakdown.json"

SECTION_RE = re.compile(r"^\s*★\s*(.+)\s*$")


def norm_district(name: str | None) -> str | None:
    if name is None:
        return None
    s = str(name).strip()
    low = s.lower()
    if low in ("city-wide", "citywide", "city wide"):
        return "citywide"
    if low == "north":
        return "north"
    if low == "south":
        return "south"
    if low == "east":
        return "east"
    if low == "west":
        return "west"
    return None


def pct_vs_2025(cur: float | int | None, prev: float | int | None) -> float | None:
    if cur is None or prev is None:
        return None
    try:
        c = float(cur)
        p = float(prev)
    except (TypeError, ValueError):
        return None
    if p == 0:
        return None
    return (c - p) / p * 100.0


def parse_sheet(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Crime Breakdown" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet 'Crime Breakdown' not found. Sheets: {wb.sheetnames}")
    ws = wb["Crime Breakdown"]
    meta_updated = None
    for r in range(1, 5):
        v = ws.cell(r, 1).value
        if v and "Last Updated" in str(v):
            meta_updated = str(v)
            break
    categories: list[dict] = []
    current: dict | None = None
    expect_header = False

    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if c1 is None or str(c1).strip() == "":
            expect_header = False
            continue
        s1 = str(c1).strip()
        if s1.startswith("*") and "Green" in s1:
            break
        m = SECTION_RE.match(s1)
        if m:
            slug = re.sub(r"[^a-z0-9]+", "-", m.group(1).lower()).strip("-")
            current = {
                "id": slug or f"cat-{len(categories)}",
                "label": m.group(1).strip(),
                "citywide": {},
                "districts": {},
            }
            categories.append(current)
            expect_header = True
            continue
        if current is None:
            continue
        if s1 == "District" and expect_header:
            expect_header = False
            continue
        if s1 == "District":
            continue

        dkey = norm_district(c1)
        if dkey is None:
            continue
        y2026 = ws.cell(r, 2).value
        y2025 = ws.cell(r, 3).value
        try:
            v26 = int(y2026) if y2026 is not None and str(y2026).strip() != "" else None
        except (TypeError, ValueError):
            v26 = None
        try:
            v25 = int(y2025) if y2025 is not None and str(y2025).strip() != "" else None
        except (TypeError, ValueError):
            v25 = None
        row = {
            "ytd2026": v26,
            "ytd2025": v25,
            "pctVs2025": pct_vs_2025(v26, v25),
        }
        if dkey == "citywide":
            current["citywide"] = row
        else:
            current["districts"][dkey] = row

    wb.close()
    return {
        "meta": {
            "sourceSheet": "Crime Breakdown",
            "sourceFile": path.name,
            "lastUpdated": meta_updated or "",
            "note": "2026 YTD vs 2025 same period — from JCPD CompStat export",
        },
        "categories": categories,
    }


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not src.is_file():
        print(f"ERROR: File not found: {src}", file=sys.stderr)
        sys.exit(1)
    data = parse_sheet(src)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"Wrote {OUT_PATH} ({len(data['categories'])} categories)")


if __name__ == "__main__":
    main()
